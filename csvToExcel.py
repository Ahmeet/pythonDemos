import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# Load the CSV file
df = pd.read_csv('ISO 9001 Ölçme Anket.csv')

# Create a new Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active
ws.title = "Quiz Results"

# Add the DataFrame to the Excel worksheet
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Apply styles
# Bold headers
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow

# Auto-adjust column width
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length

# Apply filters and freeze the first row
ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = ws["A2"]

# Optionally, add a table
tab = Table(displayName="QuizResults", ref=ws.dimensions)
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style
ws.add_table(tab)

# Save the Excel file
wb.save('beautiful_quiz_results.xlsx')
